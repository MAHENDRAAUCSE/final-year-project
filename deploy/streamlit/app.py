from __future__ import annotations

import sys
from pathlib import Path

# Add project root to Python path so src/ and other modules are discoverable
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

import io
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from inference import (
    available_targets,
    build_residual_analysis,
    load_metrics,
    load_report_text,
    predict_future,
    resolve_artifact_paths,
)


def render_residual_plot(residual_data: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(12, 4))
    ax.plot(residual_data["Index"], residual_data["Residual"], color="#ff6b6b", linewidth=1.5)
    ax.axhline(0, color="#ffffff", linestyle="--", linewidth=1)
    ax.set_title("Residuals Plot")
    ax.set_xlabel("Sample Index")
    ax.set_ylabel("Residual (Actual - Predicted)")
    ax.grid(True, alpha=0.25)
    st.pyplot(fig, width="stretch")
    plt.close(fig)


def render_prediction_plot(result_df: pd.DataFrame, target_name: str) -> None:
    """Create a professional prediction visualization."""
    fig, ax = plt.subplots(figsize=(12, 5))
    value_col = result_df.columns[1]
    
    ax.plot(result_df["Day"], result_df[value_col], marker="o", linewidth=2.5, 
            markersize=7, color="#4CAF50", label=target_name)
    ax.fill_between(result_df["Day"], result_df[value_col], alpha=0.2, color="#4CAF50")
    
    ax.set_title(f"7-Day Forecast: {target_name}", fontsize=14, fontweight="bold")
    ax.set_xlabel("Day", fontsize=11)
    ax.set_ylabel("Predicted Value (mg/L)", fontsize=11)
    ax.grid(True, alpha=0.3, linestyle="--")
    ax.legend(fontsize=10)
    
    st.pyplot(fig, width="stretch")
    plt.close(fig)


def create_excel_workbook(predictions: dict, all_targets: bool = False) -> bytes:
    """Create a formatted Excel workbook with predictions."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    if all_targets:
        # Summary sheet
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary.append(["Smart STP Predictor - All Targets Forecast"])
        ws_summary.append([])
        
        summary_data = []
        for target in predictions.keys():
            result_df = predictions[target][0]
            value_col = result_df.columns[1]
            summary_data.append({
                "Target": target,
                "Mean": f"{result_df[value_col].mean():.3f}",
                "Min": f"{result_df[value_col].min():.3f}",
                "Max": f"{result_df[value_col].max():.3f}",
                "Day 1": f"{result_df[value_col].iloc[0]:.3f}",
                "Day 7": f"{result_df[value_col].iloc[-1]:.3f}",
            })
        
        summary_df = pd.DataFrame(summary_data)
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:  # Header row
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        for col in ws_summary.columns:
            ws_summary.column_dimensions[col[0].column_letter].width = 15
        
        # Individual sheets for each target
        for target in predictions.keys():
            result_df, model_path = predictions[target]
            ws = wb.create_sheet(target.split()[0])  # Use first word as sheet name
            
            # Header
            ws.append([f"Forecast: {target}"])
            ws.append([f"Model: {model_path}"])
            ws.append([])
            
            # Data
            for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 4):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 4:  # Header
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    else:
                        if c_idx > 1:  # Value columns
                            cell.number_format = "0.000"
                        cell.alignment = Alignment(horizontal="center")
                    
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    cell.border = thin_border
            
            # Adjust widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 20
    else:
        # Single target sheet
        target = list(predictions.keys())[0]
        result_df, model_path = predictions[target]
        ws = wb.create_sheet("Forecast")
        
        # Header
        header_cell = ws.append([f"Smart STP Predictor - {target}"])[0]
        ws.append([f"Model: {model_path}"])
        ws.append([])
        
        # Statistics
        value_col = result_df.columns[1]
        stats_rows = [
            ["Statistic", "Value"],
            ["Mean", f"{result_df[value_col].mean():.3f}"],
            ["Minimum", f"{result_df[value_col].min():.3f}"],
            ["Maximum", f"{result_df[value_col].max():.3f}"],
            ["Day 1 Forecast", f"{result_df[value_col].iloc[0]:.3f}"],
            ["Day 7 Forecast", f"{result_df[value_col].iloc[-1]:.3f}"],
        ]
        
        for row in stats_rows:
            ws.append(row)
        
        ws.append([])
        ws.append(["Detailed Predictions"])
        
        # Data table
        for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 11):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 11:  # Header
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                else:
                    if c_idx > 1:
                        cell.number_format = "0.000"
                    cell.alignment = Alignment(horizontal="center")
                
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
                cell.border = thin_border
        
        # Format stats section
        for row_idx in range(4, 10):
            for col_idx in range(1, 3):
                cell = ws.cell(row=row_idx, column=col_idx)
                if col_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                if col_idx == 2:
                    cell.number_format = "0.000"
        
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


st.set_page_config(
    page_title="Smart STP Predictor",
    page_icon="💧",
    layout="wide",
)

st.title("Smart STP Predictor")
st.caption("Streamlit deployment for wastewater quality forecasting")

with st.sidebar:
    st.header("Prediction Settings")
    targets = available_targets()
    if not targets:
        st.error("No trained model files were found. Add model files in the project model paths.")
        st.stop()

    predict_all = st.checkbox("Predict All Targets", value=False)
    if not predict_all:
        selected_target = st.selectbox("Target Parameter", options=targets, index=0)
    else:
        selected_target = None
    
    days = st.slider("Future Days", min_value=1, max_value=30, value=7)
    time_steps = st.slider("Time Window", min_value=7, max_value=90, value=30)

if not predict_all and selected_target:
    metrics, metrics_path = load_metrics(selected_target)
    report_text, report_path = load_report_text(selected_target)
    artifact_paths = resolve_artifact_paths(selected_target)
else:
    metrics, metrics_path = None, None
    report_text, report_path = None, None
    artifact_paths = None

st.subheader("Upload Dataset")
st.write("Upload a CSV with historical STP data. Include the selected target column and numeric feature columns.")

uploaded_file = st.file_uploader("CSV file", type=["csv"])
df = None
residual_data = None
residual_metrics = None
residual_model_path = None
residual_error = None

if uploaded_file is not None:
    try:
        df = pd.read_csv(uploaded_file)
    except Exception as exc:
        st.error(f"Failed to read CSV file: {exc}")
        st.stop()

    st.markdown("#### Data Preview")
    st.dataframe(df.head(20), width="stretch")

    if not predict_all and selected_target:
        try:
            residual_data, residual_metrics, residual_model_path = build_residual_analysis(
                data=df,
                target_name=selected_target,
                time_steps=time_steps,
            )
        except Exception as exc:
            residual_error = str(exc)

if not predict_all:
    st.subheader("Model Evaluation")
    if metrics:
        metric_a, metric_b, metric_c, metric_d = st.columns(4)
        rmse_value = float(metrics.get("RMSE", 0.0))
        r2_value = float(metrics.get("R2 Score", metrics.get("R² Score", 0.0)))
        mae_value = float(metrics.get("MAE", 0.0))
        mape_value = float(metrics.get("MAPE (%)", 0.0))
        accuracy_value = max(0.0, 100.0 - mape_value)

        metric_a.metric("RMSE", f"{rmse_value:.3f}")
        metric_b.metric("R² Score", f"{r2_value:.3f}")
        metric_c.metric("MAE", f"{mae_value:.3f}")
        metric_d.metric("Approx. Accuracy", f"{accuracy_value:.2f}%")

        st.caption("Regression models do not have a true classification-style accuracy. The accuracy shown here is derived from 100 - MAPE.")
        if metrics_path:
            st.caption(f"Metrics source: {metrics_path}")
    else:
        st.info("No saved metrics were found for this target.")

if df is not None and not predict_all:
    st.markdown("#### Live Residual Analysis")
    if residual_data is not None and residual_metrics is not None:
        live_a, live_b, live_c, live_d = st.columns(4)
        live_a.metric("RMSE", f"{residual_metrics['RMSE']:.3f}")
        live_b.metric("R² Score", f"{residual_metrics['R2 Score']:.3f}")
        live_c.metric("MAE", f"{residual_metrics['MAE']:.3f}")
        live_d.metric("Approx. Accuracy", f"{max(0.0, 100.0 - residual_metrics['MAPE (%)']):.2f}%")

        render_residual_plot(residual_data)
        st.dataframe(residual_data.tail(20), width="stretch")
        st.caption(f"Residual analysis model: {residual_model_path}")
    else:
        st.warning(f"Residual plot could not be generated: {residual_error}")

if not predict_all and artifact_paths:
    evaluation_tabs = st.tabs(["Training History", "Predictions vs Actual", "Residuals", "Scatter", "Report"])

    with evaluation_tabs[0]:
        if artifact_paths["training_history"].exists():
            st.image(str(artifact_paths["training_history"]), width="stretch")
        else:
            st.info("Training history plot not available for this target.")

    with evaluation_tabs[1]:
        if artifact_paths["predictions"].exists():
            st.image(str(artifact_paths["predictions"]), width="stretch")
        else:
            st.info("Predictions-vs-actual plot not available for this target.")

    with evaluation_tabs[2]:
        if df is not None and residual_data is not None:
            render_residual_plot(residual_data)
        elif artifact_paths["residuals"].exists():
            st.image(str(artifact_paths["residuals"]), width="stretch")
        else:
            st.info("Residual plot not available for this target.")

    with evaluation_tabs[3]:
        if artifact_paths["scatter"].exists():
            st.image(str(artifact_paths["scatter"]), width="stretch")
        else:
            st.info("Scatter plot not available for this target.")

    with evaluation_tabs[4]:
        if report_text:
            st.text_area("Detailed report", value=report_text, height=400)
            if report_path:
                st.caption(f"Report source: {report_path}")
        else:
            st.info("Detailed report not available for this target.")

if df is not None:
    if predict_all:
        if st.button("Run Predictions for All Targets", type="primary"):
            all_results = {}
            all_success = True
            
            for target in targets:
                try:
                    result_df, model_path = predict_future(
                        data=df,
                        target_name=target,
                        days=days,
                        time_steps=time_steps,
                    )
                    all_results[target] = (result_df, model_path)
                except Exception as exc:
                    st.error(f"Prediction failed for {target}: {exc}")
                    all_success = False
            
            if all_success and all_results:
                st.success("✅ Predictions completed successfully for all targets!")
                st.markdown("---")
                
                # Display results in tabs
                tabs = st.tabs(list(all_results.keys()))
                for tab, target in zip(tabs, all_results.keys()):
                    with tab:
                        result_df, model_path = all_results[target]
                        value_col = result_df.columns[1]
                        
                        st.subheader(f"📊 {target}")
                        st.caption(f"Model: {model_path}")
                        
                        # Metrics
                        metric_col1, metric_col2, metric_col3 = st.columns(3)
                        metric_col1.metric("Mean", f"{result_df[value_col].mean():.3f}")
                        metric_col2.metric("Min", f"{result_df[value_col].min():.3f}")
                        metric_col3.metric("Max", f"{result_df[value_col].max():.3f}")
                        
                        # Plot
                        render_prediction_plot(result_df, target)
                        
                        # Data table
                        st.markdown("**Detailed Predictions**")
                        st.dataframe(result_df, width="stretch")
                
                # Excel download with formatted data
                st.markdown("---")
                st.subheader("📥 Export Results")
                excel_bytes = create_excel_workbook(all_results, all_targets=True)
                st.download_button(
                    label="📊 Download All Predictions (Excel)",
                    data=excel_bytes,
                    file_name="stp_all_targets_predictions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    else:
        if st.button("Run Prediction", type="primary"):
            try:
                result_df, model_path = predict_future(
                    data=df,
                    target_name=selected_target,
                    days=days,
                    time_steps=time_steps,
                )
            except Exception as exc:
                st.error(f"Prediction failed: {exc}")
                st.stop()

            value_col = result_df.columns[1]
            st.success("✅ Prediction completed successfully!")
            st.caption(f"Model: {model_path}")
            st.markdown("---")

            # Metrics
            metric_col1, metric_col2, metric_col3 = st.columns(3)
            metric_col1.metric("Mean", f"{result_df[value_col].mean():.3f}")
            metric_col2.metric("Min", f"{result_df[value_col].min():.3f}")
            metric_col3.metric("Max", f"{result_df[value_col].max():.3f}")

            # Plot
            st.markdown("### 📈 Forecast Visualization")
            render_prediction_plot(result_df, selected_target)

            # Data table
            st.markdown("### 📋 Detailed Forecast Data")
            st.dataframe(result_df, width="stretch")
            
            st.markdown("---")
            st.markdown("### 📥 Export Results")
            excel_bytes = create_excel_workbook({selected_target: (result_df, model_path)}, all_targets=False)
            st.download_button(
                label="📊 Download Predictions (Excel)",
                data=excel_bytes,
                file_name="stp_future_predictions.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("Upload a CSV file to start forecasting.")
